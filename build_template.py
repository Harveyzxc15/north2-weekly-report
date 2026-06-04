#!/usr/bin/env python3
"""把北一區 6 店範本轉成北二區 7 店範本。
北二區門市（顯示順序）：永和/板橋誠品/西門/花蓮/板橋遠百/新莊宏匯/新店裕隆城
"""
from copy import copy
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter as gcl
from openpyxl.worksheet.cell_range import CellRange

SRC = 'template/北一區週報_優化.xlsx'
OUT = 'template/北二區週報_優化.xlsx'

# 北二區店名（順序＝代碼 009,025,050,055,063,064,068）
N2_NAMES = ['永和門市', '板橋誠品門市', '西門門市', '花蓮門市',
            '板橋遠百門市', '新莊宏匯門市', '新店裕隆城門市']

wb = load_workbook(SRC)


def insert_rows_fix(ws, idx, amount=1):
    """在 idx 前插入 amount 列，並修正合併儲存格與列高。"""
    merges = [(m.min_row, m.max_row, m.min_col, m.max_col)
              for m in list(ws.merged_cells.ranges) if m.min_row >= idx]
    heights = {r: ws.row_dimensions[r].height
               for r in range(idx, ws.max_row + 1)
               if ws.row_dimensions.get(r) and ws.row_dimensions[r].height}
    ws.insert_rows(idx, amount)
    for (r1, r2, c1, c2) in merges:
        try:
            ws.merged_cells.remove(CellRange(f'{gcl(c1)}{r1}:{gcl(c2)}{r2}'))
        except Exception:
            pass
    for (r1, r2, c1, c2) in merges:
        ws.merge_cells(start_row=r1+amount, start_column=c1,
                       end_row=r2+amount, end_column=c2)
    for r in list(heights):
        ws.row_dimensions[r].height = None
    for r, h in heights.items():
        ws.row_dimensions[r + amount].height = h


def copy_row_format(ws, src, dst):
    """複製 src 列格式到 dst 列（不複製值）。"""
    for c in range(1, ws.max_column + 1):
        s = ws.cell(src, c); d = ws.cell(dst, c)
        if s.has_style:
            d.font = copy(s.font); d.fill = copy(s.fill)
            d.alignment = copy(s.alignment); d.border = copy(s.border)
            d.number_format = s.number_format
    h = ws.row_dimensions.get(src)
    if h and h.height:
        ws.row_dimensions[dst].height = h.height


# ── 1. BY 四張表：每個區塊在 Total 前插一列（由下而上）──
BY_SHEETS = ['BY店 本週比較', 'BY店 月累積', 'BY店 去年同期', 'BY店 整年同期']
BY_TOTAL_ROWS = [38, 29, 20, 10]   # 4 個區塊的 Total 列（由下而上插）
for sh in BY_SHEETS:
    ws = wb[sh]
    for total_r in BY_TOTAL_ROWS:
        insert_rows_fix(ws, total_r, 1)        # 在 Total 前插一空列
        copy_row_format(ws, total_r - 1, total_r)  # 沿用上一個店列的格式

# ── 1b. 本週其他細項：5 個子表各在 Total 前插一列（由下而上）──
ws_misc = wb['BY店 本週其他細項']
for total_r in [49, 39, 29, 19, 9]:
    insert_rows_fix(ws_misc, total_r, 1)
    copy_row_format(ws_misc, total_r - 1, total_r)

# ── 2. 配件：改區名 + 湊到 7 個門市分頁 ──
# 2a. 單區combined：配件-北一區 → 配件-北二區
wb['配件-北一區'].title = '配件-北二區'
wb['配件-北一區 (匯總)'].title = '配件-北二區 (匯總)'

# 2b. 門市分頁：現有 6 張，改名前 6 個北二區門市；複製第 7 張
OLD_STORE_SHEETS = ['配件 - 士林門市', '配件 - 微風門市', '配件 - 美麗華門市',
                    '配件 - 阿波羅門市', '配件 - 大葉高島屋門市', '配件 - 羅東門市']
for old, name in zip(OLD_STORE_SHEETS, N2_NAMES[:6]):
    wb[old].title = f'配件 - {name}'
# 第 7 張：複製第 6 張（花蓮）的格式，改名為第 7 個門市
def copy_rows(ws, src_start, src_end, dst_start):
    """把 src_start~src_end 整段列（值＋格式＋列高＋合併）複製到 dst_start 起。"""
    from openpyxl.cell.cell import MergedCell
    n = src_end - src_start + 1
    dst_end = dst_start + n - 1
    # 先解除目標範圍內既有的合併，避免寫入唯讀的 MergedCell
    for mm in [m for m in list(ws.merged_cells.ranges)
               if m.min_row >= dst_start and m.max_row <= dst_end]:
        ws.merged_cells.remove(mm)
    for i in range(n):
        sr, dr = src_start + i, dst_start + i
        for c in range(1, ws.max_column + 1):
            s = ws.cell(sr, c); d = ws.cell(dr, c)
            if isinstance(d, MergedCell):
                continue
            d.value = s.value
            if s.has_style:
                d.font = copy(s.font); d.fill = copy(s.fill)
                d.alignment = copy(s.alignment); d.border = copy(s.border)
                d.number_format = s.number_format
        h = ws.row_dimensions.get(sr)
        if h and h.height:
            ws.row_dimensions[dr].height = h.height
    # 合併儲存格（落在來源範圍內 → 平移到目標）
    shift = dst_start - src_start
    for m in [(mm.min_row, mm.max_row, mm.min_col, mm.max_col)
              for mm in list(ws.merged_cells.ranges)
              if mm.min_row >= src_start and mm.max_row <= src_end]:
        r1, r2, c1, c2 = m
        ws.merge_cells(start_row=r1+shift, start_column=c1,
                       end_row=r2+shift, end_column=c2)


# ── 3. 人員銷售：改 6 個區塊店名 + 複製羅東區塊成第 7 塊 ──
ws_staff = wb['BY店 人員銷售']
STAFF_TITLE_ROWS = [1, 23, 43, 66, 92, 115]   # B 欄店名所在列
for row, name in zip(STAFF_TITLE_ROWS, N2_NAMES[:6]):
    ws_staff.cell(row, 2).value = name
# 複製羅東區塊（列 115~130）到 131 起，作為第 7 塊
copy_rows(ws_staff, 115, 130, 131)
ws_staff.cell(131, 2).value = N2_NAMES[6]   # 第 7 塊標題

# ── 4. 配件匯總：複製第 3 區段成第 4 區段（放第 7 個門市，僅左欄）──
ws_sum = wb['配件-北二區 (匯總)']
copy_rows(ws_sum, 58, 76, 77)               # 第3段(列58~76) → 第4段(列77~)
# 清掉第 4 段右欄那張卡（第 7 店只用左欄）；右欄資料區 col 18~34
from openpyxl.cell.cell import MergedCell as _MC
for r in range(77, 96):
    for c in range(18, 35):
        cell = ws_sum.cell(r, c)
        if not isinstance(cell, _MC):
            cell.value = None

# 4b. 匯總每張小卡標題改成北二區店名（填入順序：左→右、由上而下）
#    band1(20): 永和|板橋誠品  band2(39): 西門|花蓮  band3(58): 板橋遠百|新莊宏匯  band4(77): 新店裕隆城
def set_card_title(ws, row, col, name):
    old = str(ws.cell(row, col).value or '')
    parts = old.split('·')
    if len(parts) >= 2:
        parts[1] = f'  {name}  '
        ws.cell(row, col).value = '·'.join(parts)
    else:
        ws.cell(row, col).value = f'配件銷售分析  ·  {name}'

CARD_TITLES = [(20, 1, '永和門市'), (20, 18, '板橋誠品門市'),
               (39, 1, '西門門市'), (39, 18, '花蓮門市'),
               (58, 1, '板橋遠百門市'), (58, 18, '新莊宏匯門市'),
               (77, 1, '新店裕隆城門市')]
for row, col, name in CARD_TITLES:
    set_card_title(ws_sum, row, col, name)

src_ws = wb[f'配件 - {N2_NAMES[5]}']
new_ws = wb.copy_worksheet(src_ws)
new_ws.title = f'配件 - {N2_NAMES[6]}'
# 把第 7 張移到第 6 張後面（複製預設會排到最後）
order = wb.sheetnames
order.remove(new_ws.title)
idx = order.index(f'配件 - {N2_NAMES[5]}') + 1
order.insert(idx, new_ws.title)
wb._sheets.sort(key=lambda s: order.index(s.title))

wb.save(OUT)
print('已輸出', OUT)
print('分頁:', wb.sheetnames)
